import tkinter as tk
from tkinter import *
from tkinter import Tk, Button, Label, filedialog, messagebox, StringVar
from PIL import Image, ImageTk
import sqlite3
import pandas as pd
import shutil
import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook
from openpyxl.styles.numbers import FORMAT_NUMBER_00  # Format for 2 decimal places
import logging
import datetime

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(f"eft_debug_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.log"),
        logging.StreamHandler()
    ]
)

# Global data frames (these need to be populated by load functions)
eft_file_df = []  # Placeholder for the eft_file_df (to be populated from the .eft file)
billing_df = []  # Placeholder for the billing_df (to be populated from the CSV)
updated_df = []  # Placeholder for the updated DataFrame
eft_header_line = ""  # Global variable to store the header line from the .eft file

def update_status(label_var, label_widget, status):
    """Update the status message for a specific process and set the color."""
    label_var.set(status)
    if status == "Complete":
        label_widget.config(fg="#00FF00")  # Green for "Complete"
    else:
        label_widget.config(fg="#FF0000")  # Red for "Not processed"

# Function to handle the rounding logic
def round_amount(amount):
    # Round the amount according to the specified rules
    amount = int(amount)
    
    # Get the last digit of the amount
    last_digit = amount % 10
    
    # If the last digit is one of the specified values, round accordingly
    if last_digit in {4, 14, 24, 34, 44, 54, 64, 74, 84, 94}:
        # Round to the next multiple of 5
        amount = (amount // 10) * 10 + 5
    elif last_digit in {9, 19, 29, 39, 49, 59, 69, 79, 89, 99}:
        # Round to the next multiple of 10
        amount = (amount // 10) * 10 + 10

    # Return the amount padded to 11 digits
    return f"{amount:011d}"

# Load CSV file and process the DataFrame
def load_csv_file():
    """Function to load the CSV file and process it as per the instructions."""
    file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
    
    if not file_path:
        return  # If no file is selected, exit the function

    # Load the CSV into a DataFrame
    global billing_df
    billing_df = pd.read_csv(file_path)

    # Consolidate data by 'SabreCode' and calculate the sum of 'TotalDue'
    billing_df = billing_df.groupby('SabreCode', as_index=False)['TotalDue'].sum()

    # Format 'SabreCode' to have a leading zero, ensuring it's 5 characters
    billing_df['SabreCode'] = billing_df['SabreCode'].apply(lambda x: f"{str(x).zfill(7)}")

    # Adjust 'TotalDue' column (multiply by 1.15 and then by 100)
    billing_df['TotalDue'] = billing_df['TotalDue'] * 1.15 * 100

    # Round the 'TotalDue' values
    billing_df['TotalDue'] = billing_df['TotalDue'].apply(round_amount)

    # Show a message box confirming the CSV data import
    messagebox.showinfo("Success", "CSV data imported successfully!")
    update_status(csv_status, csv_status_label, "Complete")    

    # Display the updated DataFrame for debugging purposes
    print(billing_df)

    # Update the status label to reflect the successful load
    update_status(csv_status, csv_status_label, "Complete")

# Load EFT file function
def load_eft_file():
    """
    Function to load an .eft file, process it into a DataFrame, and update status indicators.
    """
    global eft_file_df, column_headings, eft_header_line  # Declare global variables for the DataFrame and column headings
    
    logging.info("========== STARTING EFT FILE LOADING ==========")
    
    # Prompt user to select an .eft file
    file_path = filedialog.askopenfilename(title="Open .eft File", filetypes=(("Text files", "*.eft"), ("All files", "*.*")))
    if not file_path:
        logging.warning("File selection canceled by user")
        return  # Exit if no file is selected

    try:
        logging.info(f"Loading EFT file: {file_path}")
        with open(file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()

        logging.info(f"Total lines in file: {len(lines)}")
        
        # Capture and analyze the header line
        eft_header_line = lines[0].rstrip('\n')
        logging.debug(f"Header line: '{eft_header_line}'")
        logging.debug(f"Header line length: {len(eft_header_line)}")
        
        # Analyze the header format using regex to identify field positions
        header_positions = []
        current_pos = 0
        for char in eft_header_line:
            if char != ' ':
                if not header_positions or current_pos > header_positions[-1][1] + 1:
                    header_positions.append([current_pos, current_pos])
                else:
                    header_positions[-1][1] = current_pos
            current_pos += 1
        
        logging.debug(f"Detected header field positions: {header_positions}")

        # Skip the first line (header) and process the remaining lines
        data_lines = lines[1:]
        logging.info(f"Processing {len(data_lines)} data lines")

        # Process each line with detailed formatting checks
        processed_data = []
        column_counts = set()  # To track unique column counts
        max_columns = 0  # To keep track of the maximum number of columns
        format_issues = []  # To track any format inconsistencies

        for line_num, line in enumerate(data_lines, start=2):  # start=2 to account for skipping the first line
            line = line.rstrip('\n')
            if not line:  # Skip empty lines
                logging.warning(f"Empty line at line number {line_num}, skipping")
                continue
                
            # Debug line content and length
            logging.debug(f"Line {line_num}: '{line}'")
            logging.debug(f"Line {line_num} length: {len(line)}")
            
            # Extract the positions of non-space characters to analyze field positioning
            line_char_positions = [(i, char) for i, char in enumerate(line) if char != ' ']
            field_boundaries = []
            current_field = []
            
            for pos, char in line_char_positions:
                if not current_field or pos == current_field[-1][0] + 1:
                    current_field.append((pos, char))
                else:
                    if current_field:
                        field_boundaries.append((current_field[0][0], current_field[-1][0]))
                    current_field = [(pos, char)]
            
            if current_field:
                field_boundaries.append((current_field[0][0], current_field[-1][0]))
            
            logging.debug(f"Line {line_num} field boundaries: {field_boundaries}")
            
            # Compare field positions with the header to detect misalignments
            for i, (field_start, field_end) in enumerate(field_boundaries):
                if i < len(header_positions):
                    header_start, header_end = header_positions[i]
                    if field_start != header_start:
                        issue = f"Line {line_num}: Field {i+1} starts at position {field_start} but header field starts at {header_start}"
                        format_issues.append(issue)
                        logging.warning(issue)
            
            # Split by double spaces for regular processing
            split_line = [item.strip() for item in line.split('  ') if item.strip()]
            processed_data.append(split_line)
            
            # Check for inconsistent column counts
            column_count = len(split_line)
            column_counts.add(column_count)
            if column_count > max_columns:
                logging.info(f"Updating max columns from {max_columns} to {column_count} at line {line_num}")
                max_columns = column_count
            
            # Check if any fields exceed expected lengths based on the format
            expected_field_lengths = [7, 1, 1, 6, 19, 20, 11, 15, 1]  # Based on your formatting
            for i, field in enumerate(split_line):
                if i < len(expected_field_lengths) and len(field) > expected_field_lengths[i]:
                    issue = f"Line {line_num}: Field {i+1} '{field}' exceeds expected length of {expected_field_lengths[i]} chars"
                    format_issues.append(issue)
                    logging.warning(issue)

        # Report on column count inconsistencies
        if len(column_counts) > 1:
            logging.warning(f"Inconsistent column counts detected: {column_counts}")
            messagebox.showwarning("Warning", f"The EFT file has inconsistent column counts: {column_counts}. This may cause data misalignment.")
        else:
            logging.info(f"Consistent column count: {list(column_counts)[0]}")

        # Report on format issues
        if format_issues:
            logging.warning(f"Detected {len(format_issues)} formatting issues")
            if len(format_issues) <= 5:  # Show only the first few issues if there are many
                issue_message = "\n".join(format_issues[:5])
                messagebox.showwarning("Format Issues Detected", f"Some format issues were detected in the file:\n\n{issue_message}\n\nSee log file for details.")
            else:
                messagebox.showwarning("Format Issues Detected", f"{len(format_issues)} format issues were detected. See log file for details.")
        else:
            logging.info("No formatting issues detected")

        # Normalize the rows to match the maximum number of columns (pad shorter rows)
        for i in range(len(processed_data)):
            current_length = len(processed_data[i])
            if current_length < max_columns:
                logging.info(f"Row {i+2}: Padding row from {current_length} columns to {max_columns} columns")
                processed_data[i].extend([''] * (max_columns - current_length))  # Pad with empty strings
            elif current_length > max_columns:
                logging.info(f"Row {i+2}: Trimming row from {current_length} columns to {max_columns} columns")
                processed_data[i] = processed_data[i][:max_columns]  # Trim extra columns

        # Generate column headings dynamically
        column_headings = [f"Column {i+1}" for i in range(max_columns)]

        # Rename specific columns based on position
        if max_columns >= 1: column_headings[0] = "SabreCode"
        if max_columns >= 4: column_headings[3] = "BranchCode"
        if max_columns >= 5: column_headings[4] = "AccNumber"
        if max_columns >= 6: column_headings[5] = "CompanyName"
        if max_columns >= 7: column_headings[6] = "TotalDue"

        # Create a DataFrame with the processed data and column headings
        eft_file_df = pd.DataFrame(processed_data, columns=column_headings)

        # Log the first few rows for verification
        logging.debug("DataFrame first 5 rows:")
        for idx, row in eft_file_df.head().iterrows():
            logging.debug(f"Row {idx+1}: {dict(row)}")

        # Show a success message and update status
        logging.info("EFT File imported successfully!")
        messagebox.showinfo("Success", "EFT File imported successfully!")
        update_status(eft_status, eft_status_label, "Complete")

    except Exception as e:
        # Show an error message in case of failure
        error_message = f"An error occurred while processing the EFT file: {str(e)}"
        logging.error(error_message, exc_info=True)
        messagebox.showerror("Error", error_message)
        update_status(eft_status, eft_status_label, "Failed")
    
    finally:
        logging.info("========== COMPLETED EFT FILE LOADING ==========")

# Update Data function
def update_data():
    """
    Function to create 'updated_df' by copying 'eft_file_df' and updating the 'TotalDue'
    using values from 'billing_df' matching on 'SabreCode'. If no match exists, 'TotalDue' is set to 0.
    """
    global eft_file_df, billing_df, updated_df
    
    if eft_file_df is None or billing_df is None:
        messagebox.showerror("Error", "Please load both the EFT and Billing files before updating data.")
        return
    
    try:
        # Copy eft_file_df to create updated_df
        updated_df = eft_file_df.copy()

        # Update 'TotalDue' in updated_df by matching 'SabreCode' from billing_df
        updated_df = updated_df.merge(billing_df[['SabreCode', 'TotalDue']], on='SabreCode', how='left', suffixes=('', '_billing'))

        # If 'TotalDue_billing' is NaN, it means there was no match, so set 'TotalDue' to 0
        updated_df['TotalDue'] = updated_df['TotalDue_billing'].fillna(0)  # Replace NaN with 0 where no match was found
        
        # Drop the 'TotalDue_billing' column which we no longer need
        updated_df.drop(columns=['TotalDue_billing'], inplace=True)

        # Show a success message
        messagebox.showinfo("Info", "Updated Data created successfully!")
        update_status(updated_status, updated_status_label, "Complete")

        # Print updated_df for debugging to ensure it's correct
        print("Updated Data:")
        print(updated_df.head())

        # Optionally, if you want to display the first few rows in the Tkinter window, you can update the GUI text here.

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while updating the data: {str(e)}")
        update_status(updated_status, updated_status_label, "Failed")

# Export to Excel file function
def export_to_excel(eft_file_df, updated_df):
    # Prompt the user for the file save location
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

    if not file_path:
        return  # If no file path is selected, do nothing

    try:
        # Prepare the data for export
        updated_df_numeric = updated_df.copy()
        # Convert "TotalDue" to numeric and divide by 100
        updated_df_numeric["TotalDue"] = pd.to_numeric(updated_df_numeric["TotalDue"], errors="coerce") / 100

        # Ensure unique index for mapping (based on "SabreCode")
        updated_df_unique = updated_df_numeric.drop_duplicates(subset="SabreCode")

        # Create the export DataFrame by selecting necessary columns from eft_file_df
        export_df = eft_file_df[["SabreCode", "BranchCode", "AccNumber", "CompanyName"]].copy()

        # Map the "TotalDue" from updated_df to export_df based on "SabreCode"
        export_df["TotalDue"] = export_df["SabreCode"].map(updated_df_unique.set_index("SabreCode")["TotalDue"])

        # Ensure "eft_file_df" has unique "SabreCode" values for reindexing
        eft_file_unique = eft_file_df.drop_duplicates(subset="SabreCode")

        # Map the "TotalDue" from eft_file_df to "PrevMonthTotalDue"
        export_df["PrevMonthTotalDue"] = export_df["SabreCode"].map(eft_file_unique.set_index("SabreCode")["TotalDue"])

        # Ensure that both "TotalDue" and "PrevMonthTotalDue" are numeric (float) values
        export_df["TotalDue"] = pd.to_numeric(export_df["TotalDue"], errors="coerce")
        export_df["PrevMonthTotalDue"] = pd.to_numeric(export_df["PrevMonthTotalDue"], errors="coerce") / 100

        # Calculate the "Difference" column
        export_df["Difference"] = export_df["TotalDue"] - export_df["PrevMonthTotalDue"]

        # Now we should have exactly 7 columns: "SabreCode", "BranchCode", "AccNumber", "CompanyName", "TotalDue", "PrevMonthTotalDue", "Difference"
        column_headings = ["SabreCode", "BranchCode", "AccNumber", "CompanyName", "TotalDue", "PrevMonthTotalDue", "Difference"]

        # Create a new Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Exported Data"

        # Write the column headings with formatting
        fill_color = PatternFill(start_color="CAF2F0", end_color="CAF2F0", fill_type="solid")
        for col_num, heading in enumerate(column_headings, start=1):
            cell = ws.cell(row=1, column=col_num, value=heading)
            cell.fill = fill_color
            cell.font = Font(bold=True)

        # Write the data to the sheet
        for row_num, row_data in enumerate(export_df.itertuples(index=False), start=2):
            for col_num, (col_name, cell_value) in enumerate(zip(export_df.columns, row_data), start=1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)

                # Format numeric columns
                if col_name in ["TotalDue", "PrevMonthTotalDue", "Difference"]:
                    cell.number_format = FORMAT_NUMBER_00

                # Apply conditional formatting to the "Difference" column
                if col_name == "Difference":
                    if cell_value < 0:
                        cell.font = Font(color="FF0000")  # Red for negative values
                    elif cell_value > 0:
                        cell.font = Font(color="0000FF")  # Blue for positive values

        # Save the workbook
        wb.save(file_path)

        # Show a success message
        messagebox.showinfo("Success", "Data exported successfully!")
        update_status(export_status, export_status_label, "Complete")

    except Exception as e:
        # Show an error message if an exception occurs
        messagebox.showerror("Error", f"An error occurred while exporting: {str(e)}")

# Create new EFT file function
def create_new_eft_file():
    global updated_df, eft_header_line
    
    logging.info("========== STARTING NEW EFT FILE CREATION ==========")
    
    # Check if an EFT file has been loaded
    if not eft_header_line:
        error_msg = "No EFT file has been loaded. Please load an EFT file first to get the header format."
        logging.error(error_msg)
        messagebox.showerror("Error", error_msg)
        return
    
    # Ask the user to save the new EFT file
    save_path = filedialog.asksaveasfilename(title="Save New EFT File", defaultextension=".eft", filetypes=(("Text files", "*.eft"), ("All files", "*.*")))
    if not save_path:
        logging.info("User canceled saving new EFT file")
        return

    try:
        logging.info(f"Creating new EFT file at: {save_path}")
        logging.debug(f"Using header line: '{eft_header_line}'")
        
        # Expected field widths based on the formatting specifications
        expected_widths = [7, 1, 1, 6, 19, 20, 11, 15, 1]
        field_names = ["SabreCode", "Col2", "Col3", "BranchCode", "AccNumber", "CompanyName", "TotalDue", "SabreRadio", "NValue"]
        
        logging.debug("Expected field specifications:")
        for i, (name, width) in enumerate(zip(field_names, expected_widths)):
            logging.debug(f"  Field {i+1}: {name} - Width: {width}")
        
        # Verify updated_df structure before writing
        if updated_df is None or len(updated_df) == 0:
            error_msg = "No data available in the updated DataFrame. Please ensure data is loaded and updated first."
            logging.error(error_msg)
            messagebox.showerror("Error", error_msg)
            return
            
        logging.debug(f"DataFrame shape: {updated_df.shape}")
        
        # Verify we have all required columns
        minimum_columns = min(len(expected_widths), updated_df.shape[1])
        logging.debug(f"Minimum columns available: {minimum_columns}")
        
        # Create a counter for lines with formatting issues
        format_issues_count = 0
        format_issues = []
        
        # Open file for writing
        with open(save_path, 'w', encoding='utf-8') as new_file:
            # Write the header line
            new_file.write(eft_header_line + '\n')
            logging.debug("Header line written to file")
            
            # Process each row in the updated DataFrame
            for idx, row in updated_df.iterrows():
                # Extract values for each field
                sabre_code = str(row[0]).strip() if len(row) > 0 else ""
                col2 = str(row[1]).strip() if len(row) > 1 else ""
                col3 = str(row[2]).strip() if len(row) > 2 else ""
                branch_code = str(row[3]).strip() if len(row) > 3 else ""
                acc_number = str(row[4]).strip() if len(row) > 4 else ""
                company_name = str(row[5]).strip() if len(row) > 5 else ""
                
                # Special handling for TotalDue field when it's 0
                total_due = str(row[6]).strip() if len(row) > 6 else ""
                if total_due == "0" or total_due == "":
                    total_due = "00000000000"
                    logging.debug(f"Row {idx+1}: TotalDue was 0 or empty, replaced with '00000000000'")
                    
                sabre_radio = str(row[7]).strip() if len(row) > 7 else "SABRE RADIO"
                n_value = str(row[8]).strip() if len(row) > 8 else "N"

                # Check field lengths before formatting
                field_values = [sabre_code, col2, col3, branch_code, acc_number, company_name, total_due, sabre_radio, n_value]
                
                # Log field lengths for debugging
                field_lengths = [len(val) for val in field_values]
                logging.debug(f"Row {idx+1} field lengths: {field_lengths}")
                
                # Check if any field exceeds its expected width
                has_field_issues = False
                field_issues = []
                for i, (field_name, value, expected_width) in enumerate(zip(field_names, field_values, expected_widths)):
                    if len(value) > expected_width:
                        issue = f"Field '{field_name}' value '{value}' exceeds max width {expected_width}"
                        field_issues.append(issue)
                        has_field_issues = True
                        logging.warning(f"Row {idx+1}: {issue}")
                
                if has_field_issues:
                    format_issues_count += 1
                    # Only store the first few issues to avoid overwhelming the log
                    if len(format_issues) < 5:
                        format_issues.append(f"Row {idx+1} has formatting issues: {', '.join(field_issues)}")
                
                # Format the line with exact spacing as required
                formatted_line = (
                    f"{sabre_code:<7}  "              # SabreCode (7 chars, left-aligned) + 2 spaces
                    f"{col2:<1}  "                    # Col2 (1 char) + 2 spaces
                    f"{col3:<1}  "                    # Col3 (1 char) + 2 spaces
                    f"{branch_code:<6}  "             # BranchCode (6 chars) + 2 spaces
                    f"{acc_number:<19}  "             # AccNumber (19 chars) + 2 spaces
                    f"{company_name:<20}  "           # CompanyName (20 chars, left-aligned) + 2 spaces
                    f"{total_due:<11}  "              # TotalDue (11 chars) + 2 spaces
                    f"{sabre_radio:<15}  "            # SabreRadio (15 chars, left-aligned) + 2 spaces
                    f"{n_value}"                      # N (1 char)
                )
                
                # Verify the formatted line length and structure
                expected_formatted_length = sum(expected_widths) + 2 * (len(expected_widths) - 1)  # Add 2 spaces between each field
                actual_length = len(formatted_line)
                
                if actual_length != expected_formatted_length:
                    logging.warning(f"Row {idx+1}: Formatted line length {actual_length} doesn't match expected length {expected_formatted_length}")
                
                # Analyze spacing between fields
                fields = []
                current_field = ""
                space_count = 0
                in_space = False
                
                for char in formatted_line:
                    if (char == ' '):
                        if not in_space:
                            if current_field:
                                fields.append(current_field)
                                current_field = ""
                            in_space = True
                        space_count += 1
                    else:
                        if in_space and space_count == 2:  # We expect double spaces between fields
                            pass  # This is correct
                        elif in_space:
                            logging.warning(f"Row {idx+1}: Unexpected space count {space_count} between fields")
                            
                        in_space = False
                        space_count = 0
                        current_field += char
                
                # Add the last field if exists
                if current_field:
                    fields.append(current_field)
                
                logging.debug(f"Row {idx+1}: Parsed {len(fields)} fields from formatted line")

                # Write the formatted line to file
                new_file.write(formatted_line + '\n')
                
            # Write summary to log
            logging.info(f"Successfully processed {len(updated_df)} rows")
            if format_issues_count > 0:
                logging.warning(f"Found {format_issues_count} rows with formatting issues")
                for issue in format_issues:
                    logging.warning(issue)
            
        # Verify the output file
        logging.info(f"Verifying output file: {save_path}")
        try:
            with open(save_path, 'r', encoding='utf-8') as verify_file:
                lines = verify_file.readlines()
                logging.info(f"Output file contains {len(lines)} lines (including header)")
                
                # Verify header line
                if lines and lines[0].rstrip('\n') != eft_header_line:
                    logging.warning("Header line in output file doesn't match expected header")
                
                # Check a sample of lines for format consistency
                sample_size = min(5, len(lines) - 1)  # Check up to 5 lines (excluding header)
                for i in range(sample_size):
                    line_idx = i + 1  # Skip header line
                    line = lines[line_idx].rstrip('\n')
                    logging.debug(f"Verification sample line {line_idx}: '{line}'")
                    logging.debug(f"Sample line length: {len(line)}")
                    
                    # Check field spacing
                    space_positions = [m.start() for m in re.finditer('  ', line)]  # Find positions of double spaces
                    logging.debug(f"Double space positions: {space_positions}")
                    
                    # Verify that fields are properly separated
                    fields = [line[0:space_positions[0]]]
                    for j in range(len(space_positions)-1):
                        fields.append(line[space_positions[j]+2:space_positions[j+1]])
                    fields.append(line[space_positions[-1]+2:])
                    
                    logging.debug(f"Field count: {len(fields)}")
                    
                    # Check if we have the expected number of fields
                    if len(fields) != len(expected_widths):
                        logging.warning(f"Sample line {line_idx} has {len(fields)} fields, expected {len(expected_widths)}")
                
        except Exception as e:
            logging.error(f"Error verifying output file: {str(e)}")
        
        # Show a success message
        success_msg = "New EFT file created successfully!"
        logging.info(success_msg)
        
        if format_issues_count > 0:
            messagebox.showinfo("Success with Warnings", 
                               f"{success_msg}\n\nNote: {format_issues_count} rows had formatting issues. See log file for details.")
        else:
            messagebox.showinfo("Success", success_msg)
            
        update_status(eft_creation_status, eft_creation_status_label, "Complete")

    except Exception as e:
        # Show an error message if an exception occurs
        error_msg = f"An error occurred while creating the new EFT file: {str(e)}"
        logging.error(error_msg, exc_info=True)
        messagebox.showerror("Error", error_msg)
        update_status(eft_creation_status, eft_creation_status_label, "Failed")
    
    finally:
        logging.info("========== COMPLETED NEW EFT FILE CREATION ==========")

# Create the GUI window
root = Tk()
root.title("Debit Order Updater")

# Set the window size (width x height)
root.geometry("400x500")

# Set the background color of the window
root.configure(bg="white")

# Set the window icon (make sure the file exists in your project directory)
#root.iconbitmap(r"c:\Users\ryadya\Conda\Scripts\Debit Order\bank_78392.ico")

# Load and display the logo
try:
    # Load and resize the logo image
    logo_image = Image.open(r"C:\Users\ryadya\Conda\Scripts\DebitOrder\Final\bank.png")
    logo_image = logo_image.resize((60, 60), Image.Resampling.LANCZOS)  # Use Resampling.LANCZOS directly
    logo_photo = ImageTk.PhotoImage(logo_image)

    # Create a label to display the logo and add it to the GUI
    logo_label = tk.Label(root, image=logo_photo, bg="white")  # Adjust background to match GUI
    logo_label.image = logo_photo  # Keep a reference to avoid garbage collection
    logo_label.pack(pady=(10, 30))
except Exception as e:
    messagebox.showerror("Error", f"Unable to load logo: {str(e)}")

# Status message variables
csv_status = StringVar(value="Not processed")
eft_status = StringVar(value="Not processed")
updated_status = StringVar(value="Not processed")
export_status = StringVar(value="Not processed")
eft_creation_status = StringVar(value="Not processed")

# Create a button to load the CSV file and its status label
load_csv_button = Button(root, text="Load Bill Run CSV File", command=load_csv_file, bg="#009688", fg="white")
load_csv_button.pack(ipadx=25, pady=10)
csv_status_label = Label(root, textvariable=csv_status, bg="white", fg="#FF0000")
csv_status_label.pack(pady=2)

# Create a button to load the .eft file and its status label
load_button = Button(root, text="Load Prev. Month .eft File", command=load_eft_file, bg="#CCECFF", fg="black") # Prev color #99CCFF
load_button.pack(ipadx=16, pady=10)
eft_status_label = Label(root, textvariable=eft_status, bg="white", fg="#FF0000")
eft_status_label.pack(pady=2)

# Button to update data
update_data_button = Button(root, text="Update Data", command=update_data)
update_data_button.pack(ipadx=16, pady=10)
updated_status_label = Label(root, textvariable=eft_status, bg="white", fg="#FF0000")
updated_status_label.pack(pady=2)

# Create a button to export to Excel and its status label
export_button = Button(root, text="Export to Excel", command=lambda: export_to_excel(eft_file_df, updated_df), bg="#009688", fg="white", state="normal")
export_button.pack(ipadx=10, pady=5)
export_status_label = Label(root, textvariable=export_status, bg="white", fg="#FF0000")
export_status_label.pack(pady=2)

# Add the 'Create new EFT file' button and its status label
create_eft_button = Button(root, text="Create new EFT file", command=create_new_eft_file, bg="#66FFCC", fg="black")  # Prev color #CC0000
create_eft_button.pack(pady=5)
eft_creation_status_label = Label(root, textvariable=eft_creation_status, bg="white", fg="#FF0000")
eft_creation_status_label.pack(pady=2)

# Run the Tkinter main event loop
root.mainloop()