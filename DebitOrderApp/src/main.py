import sys
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                            QPushButton, QLabel, QFileDialog, QMessageBox,
                            QProgressBar, QHBoxLayout)
from PyQt5.QtCore import Qt, QSize
from PyQt5.QtGui import QIcon, QPixmap
import pandas as pd
import sqlite3
import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.numbers import FORMAT_NUMBER_00

class DebitOrderApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Debit Order Processor")
        self.setMinimumSize(500, 600)
        
        # Set main window style
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f8f9fa;
            }
            QWidget {
                font-family: 'Segoe UI', Arial, sans-serif;
            }
        """)
        
        # Initialize dataframes
        self.eft_file_df = None
        self.billing_df = None 
        self.updated_df = None
        
        # Create main widget and layout
        self.main_widget = QWidget()
        self.main_widget.setStyleSheet("background-color: #ffffff;")
        self.setCentralWidget(self.main_widget)
        self.layout = QVBoxLayout()
        self.layout.setContentsMargins(20, 20, 20, 20)
        self.layout.setSpacing(15)
        self.main_widget.setLayout(self.layout)
        
        # Setup UI
        self.init_ui()
        
    def init_ui(self):
        """Initialize the user interface"""
        # Add logo
        self.add_logo()
        
        # Add file loading section
        self.add_file_loading_section()
        
        # Add processing section
        self.add_processing_section()
        
        # Add export section
        self.add_export_section()
        
        # Add status bar
        self.statusBar().showMessage("Ready")
        
    def add_logo(self):
        """Add application logo"""
        logo_label = QLabel()
        pixmap = QPixmap("DebitOrderApp/resources/bank.png")
        logo_label.setPixmap(pixmap.scaled(100, 100, Qt.KeepAspectRatio))
        logo_label.setAlignment(Qt.AlignCenter)
        self.layout.addWidget(logo_label)
        
    def add_file_loading_section(self):
        """Add section for loading files"""
        # Create container widget
        file_section = QWidget()
        file_layout = QVBoxLayout()
        file_section.setLayout(file_layout)
        
        # CSV File Loading
        csv_group = QWidget()
        csv_layout = QHBoxLayout()
        csv_group.setLayout(csv_layout)
        
        self.csv_button = QPushButton("Load Bill Run CSV")
        self.csv_button.setStyleSheet("""
            QPushButton {
                background-color: #B1CBE5;
                color: white;
                border: none;
                padding: 10px 20px;
                font-size: 14px;
                border-radius: 6px;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #B1CBE5;
            }
            QPushButton:pressed {
                background-color: #B1CBE5;
            }
        """)
        self.csv_button.clicked.connect(self.load_csv_file)
        
        self.csv_status = QLabel("Not loaded")
        self.csv_status.setStyleSheet("color: #f44336;")
        
        csv_layout.addWidget(self.csv_button)
        csv_layout.addWidget(self.csv_status)
        csv_layout.addStretch()
        
        # EFT File Loading
        eft_group = QWidget()
        eft_layout = QHBoxLayout()
        eft_group.setLayout(eft_layout)
        
        self.eft_button = QPushButton("Load Previous EFT")
        self.eft_button.setStyleSheet("""
            QPushButton {
                background-color: #0984e3;
                color: white;
                border: none;
                padding: 10px 20px;
                font-size: 14px;
                border-radius: 6px;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #0767b0;
            }
            QPushButton:pressed {
                background-color: #054b80;
            }
        """)
        self.eft_button.clicked.connect(self.load_eft_file)
        
        self.eft_status = QLabel("Not loaded")
        self.eft_status.setStyleSheet("color: #f44336;")
        
        eft_layout.addWidget(self.eft_button)
        eft_layout.addWidget(self.eft_status)
        eft_layout.addStretch()
        
        # Add to main layout
        file_layout.addWidget(QLabel("<b>File Loading</b>"))
        file_layout.addWidget(csv_group)
        file_layout.addWidget(eft_group)
        self.layout.addWidget(file_section)
        
    def add_processing_section(self):
        """Add section for data processing"""
        # Create container widget
        process_section = QWidget()
        process_layout = QVBoxLayout()
        process_section.setLayout(process_layout)
        
        # Update Data button
        self.update_button = QPushButton("Update Data")
        self.update_button.setStyleSheet("""
            QPushButton {
                background-color: #0000CC;
                color: white;
                border: none;
                padding: 8px 16px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #0000CC;
            }
            QPushButton:disabled {
                background-color: #0000CC;
            }
        """)
        self.update_button.clicked.connect(self.update_data)
        self.update_button.setEnabled(False)  # Disabled until files are loaded
        
        # Status label
        self.update_status = QLabel("Not processed")
        self.update_status.setStyleSheet("color: #f44336;")
        
        # Add to layout
        process_layout.addWidget(QLabel("<b>Data Processing</b>"))
        process_layout.addWidget(self.update_button)
        process_layout.addWidget(self.update_status)
        
        self.layout.addWidget(process_section)
        
    def add_export_section(self):
        """Add section for exporting files"""
        # Create container widget
        export_section = QWidget()
        export_layout = QVBoxLayout()
        export_section.setLayout(export_layout)
        
        # Excel Export
        excel_group = QWidget()
        excel_layout = QHBoxLayout()
        excel_group.setLayout(excel_layout)
        
        self.export_button = QPushButton("Export to Excel")
        self.export_button.setStyleSheet("""
            QPushButton {
                background-color: #9BFFFF;
                color: white;
                border: none;
                padding: 10px 20px;
                font-size: 14px;
                border-radius: 6px;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #9BFFFF;
            }
            QPushButton:disabled {
                background-color: #9BFFFF;
            }
        """)
        self.export_button.clicked.connect(self.export_to_excel)
        self.export_button.setEnabled(False)
        
        self.export_status = QLabel("Not exported")
        self.export_status.setStyleSheet("color: #f44336;")
        
        excel_layout.addWidget(self.export_button)
        excel_layout.addWidget(self.export_status)
        excel_layout.addStretch()
        
        # EFT Creation
        eft_group = QWidget()
        eft_layout = QHBoxLayout()
        eft_group.setLayout(eft_layout)
        
        self.create_eft_button = QPushButton("Create New EFT")
        self.create_eft_button.setStyleSheet("""
            QPushButton {
                background-color: #607D8B;
                color: white;
                border: none;
                padding: 10px 20px;
                font-size: 14px;
                border-radius: 6px;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #455A64;
            }
            QPushButton:disabled {
                background-color: #CFD8DC;
            }
        """)
        self.create_eft_button.clicked.connect(self.create_new_eft_file)
        self.create_eft_button.setEnabled(False)
        
        self.eft_creation_status = QLabel("Not created")
        self.eft_creation_status.setStyleSheet("color: #f44336;")
        
        eft_layout.addWidget(self.create_eft_button)
        eft_layout.addWidget(self.eft_creation_status)
        eft_layout.addStretch()
        
        # Add to main layout
        export_layout.addWidget(QLabel("<b>Export Files</b>"))
        export_layout.addWidget(excel_group)
        export_layout.addWidget(eft_group)
        self.layout.addWidget(export_section)
        
    # Core functionality methods from original app
    def round_amount(self, amount):
        """Round amount according to business rules"""
        amount = int(amount)
        last_digit = amount % 10
        
        if last_digit in {4, 14, 24, 34, 44, 54, 64, 74, 84, 94}:
            amount = (amount // 10) * 10 + 5
        elif last_digit in {9, 19, 29, 39, 49, 59, 69, 79, 89, 99}:
            amount = (amount // 10) * 10 + 10
            
        return f"{amount:011d}"
        
    def load_csv_file(self):
        """Load and process CSV file"""
        try:
            # Open file dialog
            file_path, _ = QFileDialog.getOpenFileName(
                self, "Open Bill Run CSV", "", "CSV Files (*.csv)"
            )
            
            if not file_path:
                return
                
            # Show loading state
            self.csv_status.setText("Loading...")
            self.csv_status.setStyleSheet("color: #FF9800;")
            QApplication.processEvents()
            
            # Check if file has a sep=, line at the beginning
            with open(file_path, 'r', encoding='utf-8') as f:
                first_line = f.readline().strip()
                
            # Load the CSV into a DataFrame
            if (first_line.startswith('sep=')):
                # Skip the first line if it's a separator definition
                self.billing_df = pd.read_csv(file_path, skiprows=1)
            else:
                self.billing_df = pd.read_csv(file_path)
            
            print("CSV columns found:", self.billing_df.columns.tolist())
                
            # Check if the DataFrame contains CustomerCode instead of SabreCode
            if 'CustomerCode' in self.billing_df.columns:
                # Rename CustomerCode to SabreCode for consistency
                self.billing_df.rename(columns={'CustomerCode': 'SabreCode'}, inplace=True)
                print("Renamed 'CustomerCode' column to 'SabreCode'")
                
            # Verify SabreCode column exists
            if 'SabreCode' not in self.billing_df.columns:
                raise ValueError("Required column 'SabreCode' or 'CustomerCode' not found in CSV file")
            
            # Process the data
            self.billing_df = self.billing_df.groupby('SabreCode', as_index=False)['TotalDue'].sum()
            self.billing_df['SabreCode'] = self.billing_df['SabreCode'].apply(lambda x: f"{str(x).zfill(7)}")
            self.billing_df['TotalDue'] = self.billing_df['TotalDue'] * 1.15 * 100
            self.billing_df['TotalDue'] = self.billing_df['TotalDue'].apply(self.round_amount)
            
            # Update status
            self.csv_status.setText("Loaded")
            self.csv_status.setStyleSheet("color: #4CAF50;")
            self.statusBar().showMessage(f"CSV loaded: {os.path.basename(file_path)}", 5000)
            
            # Enable update button if EFT file is also loaded
            if self.eft_file_df is not None:
                self.update_button.setEnabled(True)
            
            QMessageBox.information(self, "Success", "CSV data imported successfully!")
            
        except Exception as e:
            self.csv_status.setText("Error")
            self.csv_status.setStyleSheet("color: #f44336;")
            QMessageBox.critical(self, "Error", f"Failed to load CSV: {str(e)}")
        
    def load_eft_file(self):
        """Load and process EFT file"""
        try:
            # Open file dialog
            file_path, _ = QFileDialog.getOpenFileName(
                self, "Open Previous EFT File", "", "EFT Files (*.eft);;All Files (*)"
            )
            
            if not file_path:
                return
                
            # Show loading state
            self.eft_status.setText("Loading...")
            self.eft_status.setStyleSheet("color: #FF9800;")
            QApplication.processEvents()
            
            # Process EFT file
            with open(file_path, 'r', encoding='utf-8') as file:
                lines = file.readlines()

            data_lines = lines[1:]  # Skip header
            processed_data = []
            column_counts = set()
            max_columns = 0

            for line in data_lines:
                line = line.strip()
                if not line:
                    continue
                split_line = [item.strip() for item in line.split('  ') if item.strip()]
                processed_data.append(split_line)
                column_counts.add(len(split_line))
                max_columns = max(max_columns, len(split_line))

            # Normalize rows
            for row in processed_data:
                if len(row) < max_columns:
                    row.extend([''] * (max_columns - len(row)))

            # Create column headings
            column_headings = [f"Column {i+1}" for i in range(max_columns)]
            if max_columns >= 1: column_headings[0] = "SabreCode"
            if max_columns >= 4: column_headings[3] = "BranchCode"
            if max_columns >= 5: column_headings[4] = "AccNumber"
            if max_columns >= 6: column_headings[5] = "CompanyName"
            if max_columns >= 7: column_headings[6] = "TotalDue"

            # Create DataFrame
            self.eft_file_df = pd.DataFrame(processed_data, columns=column_headings)

            # Update status
            self.eft_status.setText("Loaded")
            self.eft_status.setStyleSheet("color: #4CAF50;")
            self.statusBar().showMessage(f"EFT loaded: {os.path.basename(file_path)}", 5000)
            
            # Enable update button if CSV file is also loaded
            if self.billing_df is not None:
                self.update_button.setEnabled(True)
            
            QMessageBox.information(self, "Success", "EFT file imported successfully!")

        except Exception as e:
            self.eft_status.setText("Error")
            self.eft_status.setStyleSheet("color: #f44336;")
            QMessageBox.critical(self, "Error", f"Failed to load EFT file: {str(e)}")
        
    def enable_export_buttons(self, enabled):
        """Enable or disable export buttons"""
        self.export_button.setEnabled(enabled)
        self.create_eft_button.setEnabled(enabled)
        
    def update_data(self):
        """Update data by matching SabreCode"""
        try:
            # Check if both files are loaded
            if self.eft_file_df is None or self.billing_df is None:
                QMessageBox.warning(self, "Warning", "Please load both files first")
                return
                
            # Show processing state
            self.update_status.setText("Processing...")
            self.update_status.setStyleSheet("color: #FF9800;")
            self.update_button.setEnabled(False)
            QApplication.processEvents()
            
            print("Original EFT data:")
            print(self.eft_file_df.head())
            print("\nBilling data:")
            print(self.billing_df.head())
            
            # Create updated_df by copying eft_file_df
            self.updated_df = self.eft_file_df.copy()
            
            # Update TotalDue by matching SabreCode
            print("\nMerging data on SabreCode...")
            self.updated_df = self.updated_df.merge(
                self.billing_df[['SabreCode', 'TotalDue']],
                on='SabreCode',
                how='left',
                suffixes=('', '_billing')
            )
            
            print("\nMerged data before processing:")
            print(self.updated_df.head())
            
            # Handle unmatched records
            self.updated_df['TotalDue'] = self.updated_df['TotalDue_billing'].fillna(0)
            self.updated_df.drop(columns=['TotalDue_billing'], inplace=True)
            
            print("\nFinal updated data:")
            print(self.updated_df.head())
            
            # Update status
            self.update_status.setText("Complete")
            self.update_status.setStyleSheet("color: #4CAF50;")
            self.statusBar().showMessage("Data update complete", 5000)
            
            # Enable export buttons
            self.enable_export_buttons(True)
            
            QMessageBox.information(self, "Success", "Data updated successfully!")
            
        except Exception as e:
            self.update_status.setText("Error")
            self.update_status.setStyleSheet("color: #f44336;")
            QMessageBox.critical(self, "Error", f"Failed to update data: {str(e)}")
        finally:
            self.update_button.setEnabled(True)
        
    def export_to_excel(self):
        """Export data to Excel"""
        try:
            if self.updated_df is None:
                QMessageBox.warning(self, "Warning", "Please update data first")
                return
                
            # Show processing state
            self.export_status.setText("Exporting...")
            self.export_status.setStyleSheet("color: #FF9800;")
            self.export_button.setEnabled(False)
            QApplication.processEvents()
            
            print("\nExporting data to Excel...")
            print("Updated DataFrame columns:", self.updated_df.columns)
            print("Updated DataFrame head:")
            print(self.updated_df.head())
            
            # Get save location
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Save Excel File", "", "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                self.export_status.setText("Cancelled")
                self.export_status.setStyleSheet("color: #FF9800;")
                self.export_button.setEnabled(True)
                return
                
            # Prepare data for export
            print("\nSelecting columns for export...")
            required_cols = ["SabreCode", "BranchCode", "AccNumber", "CompanyName"]
            print("Available columns:", self.updated_df.columns)
            
            # Verify all required columns exist
            missing_cols = [col for col in required_cols if col not in self.updated_df.columns]
            if missing_cols:
                raise ValueError(f"Missing required columns: {missing_cols}")
            
            export_df = self.updated_df[required_cols].copy()
            print("\nBase export data:")
            print(export_df.head())
            
            print("\nProcessing numeric columns...")
            export_df["TotalDue"] = pd.to_numeric(self.updated_df["TotalDue"], errors='coerce') / 100
            export_df["PrevMonthTotalDue"] = pd.to_numeric(self.eft_file_df["TotalDue"], errors='coerce') / 100
            
            print("\nAfter numeric conversion:")
            print(export_df.head())
            
            export_df["Difference"] = export_df["TotalDue"] - export_df["PrevMonthTotalDue"]
            print("\nFinal export data with Difference:")
            print(export_df.head())
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Debit Order Data"
            
            # Write headers with styling
            headers = ["SabreCode", "BranchCode", "AccNumber", "CompanyName", 
                      "TotalDue", "PrevMonthTotalDue", "Difference"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CAF2F0", fill_type="solid")
            
            # Write data
            for row_num, row in enumerate(export_df.itertuples(), 2):
                ws.cell(row=row_num, column=1, value=row.SabreCode)
                ws.cell(row=row_num, column=2, value=row.BranchCode)
                ws.cell(row=row_num, column=3, value=row.AccNumber)
                ws.cell(row=row_num, column=4, value=row.CompanyName)
                ws.cell(row=row_num, column=5, value=row.TotalDue).number_format = FORMAT_NUMBER_00
                ws.cell(row=row_num, column=6, value=row.PrevMonthTotalDue).number_format = FORMAT_NUMBER_00
                diff_cell = ws.cell(row=row_num, column=7, value=row.Difference)
                diff_cell.number_format = FORMAT_NUMBER_00
                if row.Difference < 0:
                    diff_cell.font = Font(color="FF0000")
                elif row.Difference > 0:
                    diff_cell.font = Font(color="0000FF")
            
            # Save file
            wb.save(file_path)
            
            # Update status
            self.export_status.setText("Exported")
            self.export_status.setStyleSheet("color: #4CAF50;")
            self.statusBar().showMessage(f"Excel exported: {os.path.basename(file_path)}", 5000)
            QMessageBox.information(self, "Success", "Data exported to Excel successfully!")
            
        except Exception as e:
            self.export_status.setText("Error")
            self.export_status.setStyleSheet("color: #f44336;")
            QMessageBox.critical(self, "Error", f"Failed to export Excel: {str(e)}")
        finally:
            self.export_button.setEnabled(True)
        
    def create_new_eft_file(self):
        """Create new EFT file with proper fixed-width formatting that exactly matches the April 2024 2.eft format"""
        try:
            if self.updated_df is None:
                QMessageBox.warning(self, "Warning", "Please update data first")
                return
                
            # Show processing state
            self.eft_creation_status.setText("Creating...")
            self.eft_creation_status.setStyleSheet("color: #FF9800;")
            self.create_eft_button.setEnabled(False)
            QApplication.processEvents()
            
            # Get save location
            save_path, _ = QFileDialog.getSaveFileName(
                self, "Save New EFT File", "", "EFT Files (*.eft)"
            )
            
            if not save_path:
                self.eft_creation_status.setText("Cancelled")
                self.eft_creation_status.setStyleSheet("color: #FF9800;")
                self.create_eft_button.setEnabled(True)
                return
                
            # Get original EFT file for header
            original_path, _ = QFileDialog.getOpenFileName(
                self, "Select Original EFT File", "", "EFT Files (*.eft)"
            )
            
            if not original_path:
                self.eft_creation_status.setText("Cancelled")
                self.eft_creation_status.setStyleSheet("color: #FF9800;")
                self.create_eft_button.setEnabled(True)
                return
                
            # Read original file to get the header
            with open(original_path, 'r', encoding='utf-8') as file:
                original_lines = file.readlines()
                
            header = original_lines[0]  # Preserve the header line
            
            # Create new EFT file with exact formatting as in the April 2024 2.eft file
            with open(save_path, 'w', encoding='utf-8') as new_file:
                new_file.write(header)  # Write the header line first
                
                # Process each row in the updated DataFrame
                for idx, row in self.updated_df.iterrows():
                    # Extract values for each field using iloc to avoid FutureWarning
                    sabre_code = str(row.iloc[0]).strip() if len(row) > 0 else ""
                    col2 = str(row.iloc[1]).strip() if len(row) > 1 else ""
                    col3 = str(row.iloc[2]).strip() if len(row) > 2 else ""
                    branch_code = str(row.iloc[3]).strip() if len(row) > 3 else ""
                    acc_number = str(row.iloc[4]).strip() if len(row) > 4 else ""
                    company_name = str(row.iloc[5]).strip() if len(row) > 5 else ""
                    
                    # Special handling for TotalDue field when it's 0
                    total_due = str(row.iloc[6]).strip() if len(row) > 6 else ""
                    if total_due == "0" or total_due == "":
                        total_due = "00000000000"
                        
                    sabre_radio = str(row.iloc[7]).strip() if len(row) > 7 else "SABRE RADIO"
                    n_value = str(row.iloc[8]).strip() if len(row) > 8 else "N"
                    
                    # Format the line exactly as in the April 2024 2.eft file
                    # The April file format has each field with specific left-aligned width and proper spacing between fields
                    
                    # Format exactly as in the reference file with proper spacing between columns
                    formatted_line = (
                        f"{sabre_code:<7}  "              # SabreCode (7 chars, left-aligned) + 2 spaces
                        f"{col2:<1}  "                    # Col2 (1 char) + 2 spaces
                        f"{col3:<1}  "                    # Col3 (1 char) + 2 spaces
                        f"{branch_code:<6}  "             # BranchCode (6 chars) + 2 spaces
                        f"{acc_number:<19}  "             # AccNumber (19 chars) + 2 spaces
                        f"{company_name:<20}  "           # CompanyName (20 chars, left-aligned) + 2 spaces
                        f"{total_due:<11}  "          # TotalDue (11 chars) + 2 spaces
                        f"{sabre_radio:<15}  "        # SabreRadio (15 chars, left-aligned) + 2 spaces
                        f"{n_value}"                      # N (1 char)
                    )
                    
                    # Write the formatted line to file
                    new_file.write(formatted_line + '\n')
                
            # Update status
            self.eft_creation_status.setText("Created")
            self.eft_creation_status.setStyleSheet("color: #4CAF50;")
            self.statusBar().showMessage(f"EFT created: {os.path.basename(save_path)}", 5000)
            QMessageBox.information(self, "Success", "New EFT file created successfully!")
            
        except Exception as e:
            self.eft_creation_status.setText("Error")
            self.eft_creation_status.setStyleSheet("color: #f44336;")
            QMessageBox.critical(self, "Error", f"Failed to create EFT file: {str(e)}")
            print(f"Exception details: {e}")  # Print to console for debugging
        finally:
            self.create_eft_button.setEnabled(True)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = DebitOrderApp()
    window.show()
    sys.exit(app.exec_())
